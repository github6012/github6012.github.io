from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
from datetime import datetime
import os
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from urllib.parse import quote

app = Flask(__name__)

# 配置数据库
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'qhsf_hsd.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'qhsf-hsd-secret-key-2024'

# 文件上传配置
app.config['UPLOAD_FOLDER'] = os.path.join(basedir, 'static', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

# 检查文件扩展名的函数
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# 初始化扩展
db = SQLAlchemy(app)
migrate = Migrate(app, db)

# 模板上下文处理器
@app.context_processor
def inject_moment():
    return dict(moment=datetime.now)

@app.context_processor
def inject_contact_info():
    contact = ContactInfo.query.first()
    return dict(contact_info=contact)

# 数据库模型
class Admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Event(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False)
    location = db.Column(db.String(100), nullable=False)
    event_date = db.Column(db.DateTime, nullable=False)
    image_url = db.Column(db.String(200))
    category = db.Column(db.String(50), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('admin.id'))
    is_published = db.Column(db.Boolean, default=True)

class News(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    content = db.Column(db.Text, nullable=False)
    author = db.Column(db.String(100), nullable=False)
    publish_date = db.Column(db.DateTime, default=datetime.utcnow)
    image_url = db.Column(db.String(200))
    category = db.Column(db.String(50), nullable=False)
    created_by = db.Column(db.Integer, db.ForeignKey('admin.id'))
    is_published = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    university = db.Column(db.String(200), nullable=False)
    major = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    phone = db.Column(db.String(20))
    password_hash = db.Column(db.String(128))  # 密码字段
    join_date = db.Column(db.DateTime, default=datetime.utcnow)
    avatar_url = db.Column(db.String(200))
    bio = db.Column(db.Text)
    grade = db.Column(db.String(20))  # 年级字段
    group = db.Column(db.String(50))  # 分组字段
    is_approved = db.Column(db.Boolean, default=False)
    approved_by = db.Column(db.Integer, db.ForeignKey('admin.id'))
    approved_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

class Timeline(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.String(50), nullable=False)  # 时间显示文本，如"2022年3月"
    title = db.Column(db.String(200), nullable=False)  # 事件标题
    description = db.Column(db.Text, nullable=False)  # 事件描述
    marker_color = db.Column(db.String(20), default='primary')  # 标记颜色：primary, success, info, warning, danger
    order_index = db.Column(db.Integer, default=0)  # 排序索引
    is_published = db.Column(db.Boolean, default=True)  # 是否发布
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('admin.id'))
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Team(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)  # 姓名
    position = db.Column(db.String(100), nullable=False)  # 职位
    description = db.Column(db.Text, nullable=False)  # 个人描述
    avatar_color = db.Column(db.String(20), default='primary')  # 头像背景颜色：primary, success, info, warning, danger
    order_index = db.Column(db.Integer, default=0)  # 排序索引
    is_published = db.Column(db.Boolean, default=True)  # 是否发布
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('admin.id'))
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Partner(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)  # 合作伙伴名称
    icon_class = db.Column(db.String(100), nullable=False)  # 图标类名，如 'fab fa-microsoft'
    icon_color = db.Column(db.String(20), default='primary')  # 图标颜色：primary, success, info, warning, danger, dark
    website_url = db.Column(db.String(200))  # 官网链接
    description = db.Column(db.Text)  # 合作伙伴描述
    order_index = db.Column(db.Integer, default=0)  # 排序索引
    is_published = db.Column(db.Boolean, default=True)  # 是否发布
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('admin.id'))
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def check_password(self, password):
        if self.password_hash:
            return check_password_hash(self.password_hash, password)
        return False

class ContactInfo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), nullable=False)  # 联系邮箱
    phone = db.Column(db.String(20), nullable=False)  # 联系电话
    address = db.Column(db.String(200), nullable=False)  # 联系地址
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    updated_by = db.Column(db.Integer, db.ForeignKey('admin.id'))

class ContactMessage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)  # 发送者姓名
    email = db.Column(db.String(120), nullable=False)  # 发送者邮箱
    subject = db.Column(db.String(200), nullable=False)  # 消息主题
    message = db.Column(db.Text, nullable=False)  # 消息内容
    is_read = db.Column(db.Boolean, default=False)  # 是否已读
    created_at = db.Column(db.DateTime, default=datetime.utcnow)  # 创建时间
    replied_at = db.Column(db.DateTime)  # 回复时间
    replied_by = db.Column(db.Integer, db.ForeignKey('admin.id'))  # 回复者

# 管理员认证装饰器
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_id' not in session:
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

# 路由
@app.route('/')
def index():
    """首页"""
    recent_events = Event.query.order_by(Event.event_date.desc()).limit(3).all()
    recent_news = News.query.order_by(News.publish_date.desc()).limit(4).all()
    return render_template('index.html', events=recent_events, news=recent_news)

# 用户登录路由
@app.route('/login', methods=['GET', 'POST'])
def user_login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        # 检查是否为管理员
        admin = Admin.query.filter_by(email=email).first()
        if admin and admin.check_password(password) and admin.is_active:
            session['admin_id'] = admin.id
            session['admin_username'] = admin.username
            session['user_type'] = 'admin'
            admin.last_login = datetime.utcnow()
            db.session.commit()
            flash('管理员登录成功！', 'success')
            return redirect(url_for('admin_dashboard'))
        
        # 检查是否为普通用户（学生）
        student = Student.query.filter_by(email=email).first()
        if student and student.check_password(password) and student.is_approved:
            session['user_id'] = student.id
            session['username'] = student.name
            session['user_type'] = 'student'
            flash('登录成功！', 'success')
            return redirect(url_for('index'))
        elif student and not student.is_approved:
            flash('您的账户尚未通过审核，请等待管理员审核！', 'warning')
        else:
            flash('邮箱或密码错误！', 'error')
    
    return render_template('login.html')

# 用户退出路由
@app.route('/logout')
def user_logout():
    user_type = session.get('user_type')
    session.clear()
    
    if user_type == 'admin':
        flash('管理员已成功登出！', 'info')
        return redirect(url_for('admin_login'))
    else:
        flash('已成功登出！', 'info')
        return redirect(url_for('index'))

# 管理员登录路由
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        admin = Admin.query.filter_by(username=username).first()
        
        if admin and admin.check_password(password) and admin.is_active:
            session['admin_id'] = admin.id
            session['admin_username'] = admin.username
            admin.last_login = datetime.utcnow()
            db.session.commit()
            flash('登录成功！', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('用户名或密码错误！', 'error')
    
    return render_template('admin/login.html')

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_id', None)
    session.pop('admin_username', None)
    flash('已成功登出！', 'info')
    return redirect(url_for('admin_login'))

# 管理员后台主页
@app.route('/admin')
@login_required
def admin_dashboard():
    # 获取统计数据
    total_events = Event.query.count()
    total_news = News.query.count()
    total_students = Student.query.count()
    pending_students = Student.query.filter_by(is_approved=False).count()
    
    stats = {
        'total_events': total_events,
        'total_news': total_news,
        'total_students': total_students,
        'pending_students': pending_students
    }
    
    return render_template('admin/dashboard.html', stats=stats)

# 活动管理
@app.route('/admin/events')
@login_required
def admin_events():
    page = request.args.get('page', 1, type=int)
    events = Event.query.order_by(Event.created_at.desc()).paginate(
        page=page, per_page=10, error_out=False
    )
    return render_template('admin/events.html', events=events)

@app.route('/admin/events/create', methods=['GET', 'POST'])
@login_required
def admin_events_create():
    if request.method == 'POST':
        event = Event(
            title=request.form['title'],
            description=request.form['description'],
            event_date=datetime.strptime(request.form['event_date'], '%Y-%m-%dT%H:%M'),
            location=request.form['location'],
            category=request.form['category'],
            max_participants=int(request.form['max_participants']) if request.form['max_participants'] else None,
            image_url=request.form['image_url'] if request.form['image_url'] else None,
            created_by=session['admin_id'],
            is_published=bool(request.form.get('is_published'))
        )
        
        db.session.add(event)
        db.session.commit()
        flash('活动创建成功！', 'success')
        return redirect(url_for('admin_events'))
    
    return render_template('admin/events_form.html', event=None)

@app.route('/admin/events/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def admin_events_edit(id):
    event = Event.query.get_or_404(id)
    
    if request.method == 'POST':
        event.title = request.form['title']
        event.description = request.form['description']
        event.event_date = datetime.strptime(request.form['event_date'], '%Y-%m-%dT%H:%M')
        event.location = request.form['location']
        event.category = request.form['category']
        event.max_participants = int(request.form['max_participants']) if request.form['max_participants'] else None
        event.image_url = request.form['image_url'] if request.form['image_url'] else None
        event.is_published = bool(request.form.get('is_published'))
        
        db.session.commit()
        flash('活动更新成功！', 'success')
        return redirect(url_for('admin_events'))
    
    return render_template('admin/events_form.html', event=event)

@app.route('/admin/events/<int:id>/delete', methods=['POST'])
@login_required
def admin_events_delete(id):
    event = Event.query.get_or_404(id)
    db.session.delete(event)
    db.session.commit()
    flash('活动删除成功！', 'success')
    return redirect(url_for('admin_events'))

# 学生管理（已注册学生的信息管理）
@app.route('/admin/students')
@login_required
def admin_students():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    grade = request.args.get('grade', '')
    group = request.args.get('group', '')
    
    # 只显示已批准的学生
    query = Student.query.filter(Student.is_approved == True)
    
    if search:
        query = query.filter(
            db.or_(
                Student.name.contains(search),
                Student.university.contains(search),
                Student.major.contains(search),
                Student.email.contains(search)
            )
        )
    
    if grade:
        query = query.filter(Student.grade == grade)
        
    if group:
        query = query.filter(Student.group == group)
    
    students = query.order_by(Student.created_at.desc()).paginate(
        page=page, per_page=15, error_out=False
    )
    
    # 获取所有年级和分组用于筛选
    grades = db.session.query(Student.grade).distinct().all()
    grades = [g[0] for g in grades if g[0]]
    
    groups = db.session.query(Student.group).distinct().all()
    groups = [g[0] for g in groups if g[0]]
    
    return render_template('admin/students.html', 
                         students=students, 
                         search=search, 
                         grade=grade, 
                         group=group,
                         grades=grades,
                         groups=groups)

# 申请审核（专门用于审核新申请）
@app.route('/admin/applications')
@login_required
def admin_applications():
    page = request.args.get('page', 1, type=int)
    
    # 只显示待审核的申请
    query = Student.query.filter(Student.is_approved == False)
    
    students = query.order_by(Student.created_at.desc()).paginate(
        page=page, per_page=10, error_out=False
    )
    return render_template('admin/applications.html', students=students)

@app.route('/admin/students/<int:id>/approve', methods=['POST'])
@login_required
def admin_students_approve(id):
    student = Student.query.get_or_404(id)
    student.is_approved = True
    student.approved_by = session['admin_id']
    student.approved_at = datetime.utcnow()
    
    db.session.commit()
    flash(f'已批准 {student.name} 的申请！', 'success')
    return redirect(url_for('admin_applications'))

@app.route('/admin/students/<int:id>/reject', methods=['POST'])
@login_required
def admin_students_reject(id):
    student = Student.query.get_or_404(id)
    student.is_approved = False
    student.approved_by = None
    student.approved_at = None
    
    db.session.commit()
    flash(f'已拒绝 {student.name} 的申请！', 'warning')
    return redirect(url_for('admin_applications'))

# 学生分组管理
@app.route('/admin/students/<int:id>/group', methods=['POST'])
@login_required
def admin_students_group(id):
    student = Student.query.get_or_404(id)
    group = request.form.get('group')
    student.group = group
    
    db.session.commit()
    flash(f'已将 {student.name} 分配到 {group} 组！', 'success')
    return redirect(url_for('admin_students'))

@app.route('/admin/students/<int:id>/delete', methods=['POST'])
@login_required
def admin_students_delete(id):
    student = Student.query.get_or_404(id)
    db.session.delete(student)
    db.session.commit()
    flash('学生信息删除成功！', 'success')
    return redirect(url_for('admin_students'))

# 新闻管理
@app.route('/admin/news')
@login_required
def admin_news():
    page = request.args.get('page', 1, type=int)
    news = News.query.order_by(News.created_at.desc()).paginate(
        page=page, per_page=10, error_out=False
    )
    return render_template('admin/news.html', news=news)

@app.route('/admin/news/create', methods=['GET', 'POST'])
@login_required
def admin_news_create():
    if request.method == 'POST':
        # 处理文件上传
        image_url = None
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename != '' and allowed_file(file.filename):
                # 生成安全的文件名
                filename = secure_filename(file.filename)
                # 添加UUID前缀避免文件名冲突
                unique_filename = f"{uuid.uuid4().hex}_{filename}"
                # 确保上传目录存在
                news_upload_path = os.path.join(app.config['UPLOAD_FOLDER'], 'news')
                os.makedirs(news_upload_path, exist_ok=True)
                # 保存文件
                file_path = os.path.join(news_upload_path, unique_filename)
                file.save(file_path)
                # 设置相对URL路径
                image_url = f"/static/uploads/news/{unique_filename}"
        
        news = News(
            title=request.form['title'],
            content=request.form['content'],
            category=request.form['category'],
            image_url=image_url,
            created_by=session['admin_id'],
            is_published=bool(request.form.get('is_published'))
        )
        
        db.session.add(news)
        db.session.commit()
        flash('新闻创建成功！', 'success')
        return redirect(url_for('admin_news'))
    
    return render_template('admin/news_form.html', news=None)

@app.route('/admin/news/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def admin_news_edit(id):
    news = News.query.get_or_404(id)
    
    if request.method == 'POST':
        # 处理文件上传
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename != '' and allowed_file(file.filename):
                # 删除旧图片文件（如果存在）
                if news.image_url and news.image_url.startswith('/static/uploads/'):
                    old_file_path = os.path.join(basedir, news.image_url.lstrip('/'))
                    if os.path.exists(old_file_path):
                        os.remove(old_file_path)
                
                # 生成安全的文件名
                filename = secure_filename(file.filename)
                # 添加UUID前缀避免文件名冲突
                unique_filename = f"{uuid.uuid4().hex}_{filename}"
                # 确保上传目录存在
                news_upload_path = os.path.join(app.config['UPLOAD_FOLDER'], 'news')
                os.makedirs(news_upload_path, exist_ok=True)
                # 保存文件
                file_path = os.path.join(news_upload_path, unique_filename)
                file.save(file_path)
                # 设置相对URL路径
                news.image_url = f"/static/uploads/news/{unique_filename}"
        
        news.title = request.form['title']
        news.content = request.form['content']
        news.category = request.form['category']
        news.is_published = bool(request.form.get('is_published'))
        
        db.session.commit()
        flash('新闻更新成功！', 'success')
        return redirect(url_for('admin_news'))
    
    return render_template('admin/news_form.html', news=news)

@app.route('/admin/news/<int:id>/delete', methods=['POST'])
@login_required
def admin_news_delete(id):
    news = News.query.get_or_404(id)
    db.session.delete(news)
    db.session.commit()
    flash('新闻删除成功！', 'success')
    return redirect(url_for('admin_news'))

# 时间线管理
@app.route('/admin/timeline')
@login_required
def admin_timeline():
    page = request.args.get('page', 1, type=int)
    timelines = Timeline.query.order_by(Timeline.order_index.asc(), Timeline.created_at.desc()).paginate(
        page=page, per_page=10, error_out=False)
    return render_template('admin/timeline.html', timelines=timelines)

@app.route('/admin/timeline/create', methods=['GET', 'POST'])
@login_required
def admin_timeline_create():
    if request.method == 'POST':
        timeline = Timeline(
            date=request.form['date'],
            title=request.form['title'],
            description=request.form['description'],
            marker_color=request.form.get('marker_color', 'primary'),
            order_index=int(request.form.get('order_index', 0)),
            is_published=bool(request.form.get('is_published')),
            created_by=session['admin_id']
        )
        db.session.add(timeline)
        db.session.commit()
        flash('时间线事件创建成功！', 'success')
        return redirect(url_for('admin_timeline'))
    return render_template('admin/timeline_form.html')

@app.route('/admin/timeline/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def admin_timeline_edit(id):
    timeline = Timeline.query.get_or_404(id)
    if request.method == 'POST':
        timeline.date = request.form['date']
        timeline.title = request.form['title']
        timeline.description = request.form['description']
        timeline.marker_color = request.form.get('marker_color', 'primary')
        timeline.order_index = int(request.form.get('order_index', 0))
        timeline.is_published = bool(request.form.get('is_published'))
        timeline.updated_at = datetime.utcnow()
        db.session.commit()
        flash('时间线事件更新成功！', 'success')
        return redirect(url_for('admin_timeline'))
    return render_template('admin/timeline_form.html', timeline=timeline)

@app.route('/admin/timeline/<int:id>/delete', methods=['POST'])
@login_required
def admin_timeline_delete(id):
    timeline = Timeline.query.get_or_404(id)
    db.session.delete(timeline)
    db.session.commit()
    flash('时间线事件删除成功！', 'success')
    return redirect(url_for('admin_timeline'))

# 团队管理
@app.route('/admin/team')
@login_required
def admin_team():
    page = request.args.get('page', 1, type=int)
    teams = Team.query.order_by(Team.order_index.asc(), Team.created_at.desc()).paginate(
        page=page, per_page=10, error_out=False)
    return render_template('admin/team.html', teams=teams)

@app.route('/admin/team/create', methods=['GET', 'POST'])
@login_required
def admin_team_create():
    if request.method == 'POST':
        team = Team(
            name=request.form['name'],
            position=request.form['position'],
            description=request.form['description'],
            avatar_color=request.form.get('avatar_color', 'primary'),
            order_index=int(request.form.get('order_index', 0)),
            is_published=bool(request.form.get('is_published')),
            created_by=session['admin_id']
        )
        db.session.add(team)
        db.session.commit()
        flash('团队成员创建成功！', 'success')
        return redirect(url_for('admin_team'))
    return render_template('admin/team_form.html')

@app.route('/admin/team/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def admin_team_edit(id):
    team = Team.query.get_or_404(id)
    if request.method == 'POST':
        team.name = request.form['name']
        team.position = request.form['position']
        team.description = request.form['description']
        team.avatar_color = request.form.get('avatar_color', 'primary')
        team.order_index = int(request.form.get('order_index', 0))
        team.is_published = bool(request.form.get('is_published'))
        team.updated_at = datetime.utcnow()
        db.session.commit()
        flash('团队成员更新成功！', 'success')
        return redirect(url_for('admin_team'))
    return render_template('admin/team_form.html', team=team)

@app.route('/admin/team/<int:id>/delete', methods=['POST'])
@login_required
def admin_team_delete(id):
    team = Team.query.get_or_404(id)
    db.session.delete(team)
    db.session.commit()
    flash('团队成员删除成功！', 'success')
    return redirect(url_for('admin_team'))

# 合作伙伴管理
@app.route('/admin/partners')
@login_required
def admin_partners():
    page = request.args.get('page', 1, type=int)
    partners = Partner.query.order_by(Partner.order_index.asc(), Partner.created_at.desc()).paginate(
        page=page, per_page=12, error_out=False)
    return render_template('admin/partners.html', partners=partners)

@app.route('/admin/partners/create', methods=['GET', 'POST'])
@login_required
def admin_partners_create():
    if request.method == 'POST':
        partner = Partner(
            name=request.form['name'],
            icon_class=request.form['icon_class'],
            icon_color=request.form.get('icon_color', 'primary'),
            website_url=request.form.get('website_url', ''),
            description=request.form.get('description', ''),
            order_index=int(request.form.get('order_index', 0)),
            is_published=bool(request.form.get('is_published')),
            created_by=session['admin_id']
        )
        db.session.add(partner)
        db.session.commit()
        flash('合作伙伴创建成功！', 'success')
        return redirect(url_for('admin_partners'))
    return render_template('admin/partners_form.html')

@app.route('/admin/partners/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def admin_partners_edit(id):
    partner = Partner.query.get_or_404(id)
    if request.method == 'POST':
        partner.name = request.form['name']
        partner.icon_class = request.form['icon_class']
        partner.icon_color = request.form.get('icon_color', 'primary')
        partner.website_url = request.form.get('website_url', '')
        partner.description = request.form.get('description', '')
        partner.order_index = int(request.form.get('order_index', 0))
        partner.is_published = bool(request.form.get('is_published'))
        partner.updated_at = datetime.utcnow()
        db.session.commit()
        flash('合作伙伴更新成功！', 'success')
        return redirect(url_for('admin_partners'))
    return render_template('admin/partners_form.html', partner=partner)

@app.route('/admin/partners/<int:id>/delete', methods=['POST'])
@login_required
def admin_partners_delete(id):
    partner = Partner.query.get_or_404(id)
    db.session.delete(partner)
    db.session.commit()
    flash('合作伙伴删除成功！', 'success')
    return redirect(url_for('admin_partners'))

# 联系信息管理
@app.route('/admin/contact')
@login_required
def admin_contact():
    contact = ContactInfo.query.first()
    return render_template('admin/contact.html', contact=contact)

@app.route('/admin/contact/edit', methods=['GET', 'POST'])
@login_required
def admin_contact_edit():
    contact = ContactInfo.query.first()
    
    if request.method == 'POST':
        if contact:
            # 更新现有联系信息
            contact.email = request.form['email']
            contact.phone = request.form['phone']
            contact.address = request.form['address']
            contact.updated_at = datetime.utcnow()
            contact.updated_by = session['admin_id']
        else:
            # 创建新的联系信息
            contact = ContactInfo(
                email=request.form['email'],
                phone=request.form['phone'],
                address=request.form['address'],
                updated_by=session['admin_id']
            )
            db.session.add(contact)
        
        db.session.commit()
        flash('联系信息更新成功！', 'success')
        return redirect(url_for('admin_contact'))
    
    return render_template('admin/contact_form.html', contact=contact)

# 联系消息管理
@app.route('/admin/messages')
@login_required
def admin_messages():
    """联系消息管理页面"""
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    messages = ContactMessage.query.order_by(ContactMessage.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('admin/messages.html', messages=messages)

@app.route('/admin/messages/<int:message_id>/read', methods=['POST'])
@login_required
def admin_message_read(message_id):
    """标记消息为已读"""
    message = ContactMessage.query.get_or_404(message_id)
    message.is_read = True
    message.replied_by = session['admin_id']
    message.replied_at = datetime.utcnow()
    db.session.commit()
    
    return jsonify({'success': True, 'message': '消息已标记为已读'})

@app.route('/admin/messages/<int:message_id>/delete', methods=['POST'])
@login_required
def admin_message_delete(message_id):
    """删除消息"""
    message = ContactMessage.query.get_or_404(message_id)
    db.session.delete(message)
    db.session.commit()
    
    flash('消息删除成功！', 'success')
    return redirect(url_for('admin_messages'))

# 学生信息管理
@app.route('/admin/students/info')
@login_required
def admin_students_info():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    grade = request.args.get('grade', '')
    
    query = Student.query
    
    if search:
        query = query.filter(
            db.or_(
                Student.name.contains(search),
                Student.university.contains(search),
                Student.major.contains(search),
                Student.email.contains(search)
            )
        )
    
    if grade:
        query = query.filter(Student.grade == grade)
    
    students = query.order_by(Student.created_at.desc()).paginate(
        page=page, per_page=15, error_out=False
    )
    
    # 获取所有年级用于筛选
    grades = db.session.query(Student.grade).distinct().all()
    grades = [g[0] for g in grades if g[0]]
    
    return render_template('admin/students_info.html', 
                         students=students, 
                         search=search, 
                         grade=grade, 
                         grades=grades)

@app.route('/admin/students/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def admin_students_edit(id):
    student = Student.query.get_or_404(id)
    
    if request.method == 'POST':
        student.name = request.form['name']
        student.university = request.form['university']
        student.major = request.form['major']
        student.grade = request.form['grade']
        student.email = request.form['email']
        student.phone = request.form['phone']
        student.bio = request.form['bio']
        
        db.session.commit()
        flash('学生信息更新成功！', 'success')
        return redirect(url_for('admin_students_info'))
    
    return render_template('admin/students_edit.html', student=student)

# 导出学生信息为Excel
@app.route('/admin/students/export')
@login_required
def admin_students_export():
    # 获取所有学生信息
    students = Student.query.all()
    
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "学生信息"
    
    # 设置表头
    headers = ['ID', '姓名', '大学', '专业', '年级', '邮箱', '电话', '加入日期', '审核状态', '个人简介']
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 添加学生数据
    for student in students:
        row_data = [
            student.id,
            student.name,
            student.university,
            student.major,
            student.grade or '未设置',
            student.email,
            student.phone or '未填写',
            student.join_date.strftime('%Y-%m-%d') if student.join_date else '未知',
            '已审核' if student.is_approved else '待审核',
            student.bio or '无'
        ]
        ws.append(row_data)
    
    # 调整列宽
    column_widths = [8, 15, 25, 20, 10, 25, 15, 12, 10, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    # 创建内存中的文件
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 创建响应
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    # 使用URL编码处理中文文件名
    filename = f'学生信息_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    encoded_filename = quote(filename.encode('utf-8'))
    response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'\'{encoded_filename}'
    
    return response

# 数据统计和分析
@app.route('/admin/statistics')
@login_required
def admin_statistics():
    # 基础统计数据
    total_events = Event.query.count()
    published_events = Event.query.filter_by(is_published=True).count()
    total_news = News.query.count()
    published_news = News.query.filter_by(is_published=True).count()
    total_students = Student.query.count()
    approved_students = Student.query.filter_by(is_approved=1).count()
    pending_students = Student.query.filter(Student.is_approved.is_(None)).count()
    
    # 月度统计数据
    from datetime import datetime, timedelta
    import calendar
    
    # 获取最近12个月的数据
    monthly_data = []
    current_date = datetime.now()
    
    for i in range(12):
        # 计算月份
        month_date = current_date - timedelta(days=30*i)
        month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if month_date.month == 12:
            month_end = month_date.replace(year=month_date.year+1, month=1, day=1) - timedelta(seconds=1)
        else:
            month_end = month_date.replace(month=month_date.month+1, day=1) - timedelta(seconds=1)
        
        # 统计该月数据
        month_events = Event.query.filter(
            Event.created_at >= month_start,
            Event.created_at <= month_end
        ).count()
        
        month_news = News.query.filter(
            News.created_at >= month_start,
            News.created_at <= month_end
        ).count()
        
        month_students = Student.query.filter(
            Student.created_at >= month_start,
            Student.created_at <= month_end
        ).count()
        
        monthly_data.append({
            'month': month_date.strftime('%Y-%m'),
            'month_name': f"{month_date.year}年{month_date.month}月",
            'events': month_events,
            'news': month_news,
            'students': month_students
        })
    
    monthly_data.reverse()  # 按时间正序排列
    
    # 活动分类统计
    event_categories_raw = db.session.query(
        Event.category, 
        db.func.count(Event.id).label('count')
    ).group_by(Event.category).all()
    event_categories = [{'category': row.category, 'count': row.count} for row in event_categories_raw]
    
    # 新闻分类统计
    news_categories_raw = db.session.query(
        News.category, 
        db.func.count(News.id).label('count')
    ).group_by(News.category).all()
    news_categories = [{'category': row.category, 'count': row.count} for row in news_categories_raw]
    
    # 学生年级分布
    student_grades_raw = db.session.query(
        Student.grade, 
        db.func.count(Student.id).label('count')
    ).group_by(Student.grade).all()
    student_grades = [{'grade': row.grade, 'count': row.count} for row in student_grades_raw]
    

    
    return render_template('admin/statistics.html',
                         total_events=total_events,
                         published_events=published_events,
                         total_news=total_news,
                         published_news=published_news,
                         total_students=total_students,
                         approved_students=approved_students,
                         pending_students=pending_students,
                         monthly_data=monthly_data,
                         event_categories=event_categories,
                         news_categories=news_categories,
                         student_grades=student_grades)

@app.route('/events')
def events():
    """活动页面"""
    category = request.args.get('category', '')
    if category:
        events = Event.query.filter_by(category=category).order_by(Event.event_date.desc()).all()
    else:
        events = Event.query.order_by(Event.event_date.desc()).all()
    return render_template('events.html', events=events, current_category=category)

@app.route('/event/<int:event_id>')
def event_detail(event_id):
    """活动详情页面"""
    event = Event.query.get_or_404(event_id)
    return render_template('event_detail.html', event=event)

@app.route('/news')
def news():
    """新闻页面"""
    category = request.args.get('category', '')
    if category:
        news_list = News.query.filter_by(category=category).order_by(News.publish_date.desc()).all()
    else:
        news_list = News.query.order_by(News.publish_date.desc()).all()
    return render_template('news.html', news_list=news_list, current_category=category)

@app.route('/news/<int:news_id>')
def news_detail(news_id):
    """新闻详情页面"""
    news_item = News.query.get_or_404(news_id)
    return render_template('news_detail.html', news=news_item)

@app.route('/students')
def students():
    """学生展示页面"""
    students = Student.query.order_by(Student.join_date.desc()).all()
    return render_template('students.html', students=students)

@app.route('/join')
def join():
    """加入我们页面"""
    return render_template('join.html')

# API路由
@app.route('/api/join', methods=['POST'])
def api_join():
    """处理加入申请"""
    data = request.get_json()
    
    # 检查邮箱是否已存在
    existing_student = Student.query.filter_by(email=data['email']).first()
    if existing_student:
        return jsonify({'success': False, 'message': '该邮箱已注册'})
    
    # 检查密码是否提供
    if not data.get('password'):
        return jsonify({'success': False, 'message': '密码不能为空'})
    
    # 创建新学生记录
    new_student = Student(
        name=data['name'],
        university=data['university'],
        major=data['major'],
        email=data['email'],
        phone=data.get('phone', ''),
        bio=data.get('bio', '')
    )
    
    # 设置密码
    new_student.set_password(data['password'])
    
    try:
        db.session.add(new_student)
        db.session.commit()
        return jsonify({'success': True, 'message': '申请提交成功！'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '提交失败，请重试'})

@app.route('/api/events', methods=['GET'])
def api_events():
    """获取活动列表API"""
    category = request.args.get('category', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    
    query = Event.query
    if category:
        query = query.filter_by(category=category)
    
    events = query.order_by(Event.event_date.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'success': True,
        'events': [{
            'id': event.id,
            'title': event.title,
            'description': event.description,
            'location': event.location,
            'event_date': event.event_date.isoformat(),
            'category': event.category,
            'image_url': event.image_url
        } for event in events.items],
        'pagination': {
            'page': events.page,
            'pages': events.pages,
            'per_page': events.per_page,
            'total': events.total
        }
    })

@app.route('/api/news', methods=['GET'])
def api_news():
    """获取新闻列表API"""
    category = request.args.get('category', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    
    query = News.query
    if category:
        query = query.filter_by(category=category)
    
    news = query.order_by(News.publish_date.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'success': True,
        'news': [{
            'id': item.id,
            'title': item.title,
            'content': item.content[:200] + '...' if len(item.content) > 200 else item.content,
            'author': item.author,
            'publish_date': item.publish_date.isoformat(),
            'category': item.category,
            'image_url': item.image_url
        } for item in news.items],
        'pagination': {
            'page': news.page,
            'pages': news.pages,
            'per_page': news.per_page,
            'total': news.total
        }
    })

@app.route('/api/students', methods=['GET'])
def api_students():
    """获取学生列表API"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 12, type=int)
    
    students = Student.query.order_by(Student.join_date.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'success': True,
        'students': [{
            'id': student.id,
            'name': student.name,
            'university': student.university,
            'major': student.major,
            'join_date': student.join_date.isoformat(),
            'avatar_url': student.avatar_url,
            'bio': student.bio
        } for student in students.items],
        'pagination': {
            'page': students.page,
            'pages': students.pages,
            'per_page': students.per_page,
            'total': students.total
        }
    })

@app.route('/api/stats', methods=['GET'])
def api_stats():
    """获取统计数据API"""
    total_students = Student.query.count()
    total_events = Event.query.count()
    total_news = News.query.count()
    
    # 计算合作高校数量（基于学生的大学去重）
    universities = db.session.query(Student.university).distinct().count()
    
    return jsonify({
        'success': True,
        'stats': {
            'total_students': total_students,
            'total_events': total_events,
            'total_news': total_news,
            'universities': universities
        }
    })

@app.route('/api/subscribe', methods=['POST'])
def api_subscribe():
    """处理新闻订阅"""
    data = request.get_json()
    email = data.get('email')
    
    if not email:
        return jsonify({'success': False, 'message': '请提供有效的邮箱地址'})
    
    # 这里可以添加邮箱订阅逻辑，比如保存到数据库或发送到邮件服务
    # 目前只是简单返回成功消息
    return jsonify({'success': True, 'message': '订阅成功！我们会定期向您发送最新资讯。'})

@app.route('/api/contact', methods=['POST'])
def api_contact():
    """处理联系表单"""
    data = request.get_json()
    
    required_fields = ['name', 'email', 'subject', 'message']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'success': False, 'message': f'请填写{field}字段'})
    
    try:
        # 创建新的联系消息记录
        contact_message = ContactMessage(
            name=data['name'],
            email=data['email'],
            subject=data['subject'],
            message=data['message']
        )
        
        # 保存到数据库
        db.session.add(contact_message)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '消息发送成功！我们会尽快回复您。'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '发送失败，请重试。'})

@app.route('/about')
def about():
    """关于我们页面"""
    # 获取已发布的时间线事件，按排序索引和创建时间排序
    timelines = Timeline.query.filter_by(is_published=True).order_by(
        Timeline.order_index.asc(), Timeline.created_at.desc()).all()
    # 获取已发布的团队成员
    teams = Team.query.filter_by(is_published=True).order_by(Team.order_index.asc(), Team.created_at.desc()).all()
    # 获取已发布的合作伙伴数据，按order_index和创建时间排序
    partners = Partner.query.filter_by(is_published=True).order_by(Partner.order_index.asc(), Partner.created_at.desc()).all()
    return render_template('about.html', timelines=timelines, teams=teams, partners=partners)

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # 创建默认管理员账户（如果不存在）
        admin = Admin.query.filter_by(username='admin').first()
        if not admin:
            admin = Admin(
                username='admin',
                email='admin@qhsf-hsd.com',
                is_active=True
            )
            admin.set_password('admin123')  # 默认密码
            db.session.add(admin)
            db.session.commit()
            print('默认管理员账户已创建：用户名=admin，密码=admin123')
        

    
    app.run(debug=True, host='0.0.0.0', port=5000)